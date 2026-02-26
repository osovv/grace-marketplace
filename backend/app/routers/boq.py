import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.deps import get_db, get_current_user, require_role
from app.models import BOQItem, Project
from app.schemas import BOQItemOut, BOQItemPatch
from app.audit import log_audit

router = APIRouter(tags=["boq"])


@router.post("/api/projects/{project_id}/boq/upload", response_model=list[BOQItemOut])
async def upload_boq(
    project_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role("planner")),
    db: AsyncSession = Depends(get_db),
):
    # Verify project belongs to tenant
    result = await db.execute(
        select(Project).where(Project.id == project_id, Project.tenant_id == current_user["tenant_id"])
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    filename = file.filename or ""

    items = []
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
            for row in rows_iter:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = str(val).strip() if val is not None else ""
                if row_dict:
                    items.append(row_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {e}")
    else:
        # CSV
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                cleaned = {k.strip().lower(): v.strip() for k, v in row.items() if k}
                items.append(cleaned)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {e}")

    if not items:
        raise HTTPException(status_code=400, detail="No data rows found")

    created = []
    for row in items:
        qty_str = row.get("qty", "0")
        try:
            qty_val = float(qty_str) if qty_str else 0
        except ValueError:
            qty_val = 0

        boq = BOQItem(
            tenant_id=current_user["tenant_id"],
            project_id=project_id,
            code=row.get("code", ""),
            description=row.get("description", ""),
            qty=qty_val,
            uom=row.get("uom", ""),
            area_raw=row.get("area", ""),
            area_norm=row.get("area_norm", ""),
            discipline=row.get("discipline", ""),
            work_type=row.get("work_type", ""),
        )
        db.add(boq)
        created.append(boq)

    await db.flush()
    await log_audit(
        db, current_user["tenant_id"], current_user["user_id"],
        "boq", project_id, "IMPORT", {"count": len(created)},
    )
    await db.commit()

    for item in created:
        await db.refresh(item)

    return created


@router.get("/api/projects/{project_id}/boq", response_model=list[BOQItemOut])
async def list_boq(
    project_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BOQItem).where(
            BOQItem.project_id == project_id,
            BOQItem.tenant_id == current_user["tenant_id"],
        )
    )
    return result.scalars().all()


@router.patch("/api/boq/{boq_item_id}", response_model=BOQItemOut)
async def patch_boq_item(
    boq_item_id: str,
    patch: BOQItemPatch,
    current_user: dict = Depends(require_role("planner")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BOQItem).where(
            BOQItem.id == boq_item_id,
            BOQItem.tenant_id == current_user["tenant_id"],
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")

    changes = {}
    for field, val in patch.model_dump(exclude_unset=True).items():
        if val is not None:
            old = getattr(item, field)
            setattr(item, field, val)
            changes[field] = {"old": str(old), "new": str(val)}

    if changes:
        await log_audit(
            db, current_user["tenant_id"], current_user["user_id"],
            "boq_item", boq_item_id, "UPDATE", changes,
        )
        await db.commit()
        await db.refresh(item)

    return item
